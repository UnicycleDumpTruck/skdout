import yaml
import sys
import csv
import datetime
import pandas as pd
import openpyxl


def time_to_date(input_time):
    return datetime.datetime.combine(datetime.date(2020, 1, 1), input_time)


def prep_df(day_of_week):
    task_list = Task.objects.filter(weekday=day_of_week.lower().capitalize())
    no_virex = task_list.exclude(task_text="Virex Spray Everything")
    q = no_virex.values(
        "location",
        "start_time",
        "end_time",
        "task_text",
        "weekday",
    )

    df = pd.DataFrame.from_records(q)
    df["start_date"] = df.start_time.apply(time_to_date)
    df["end_date"] = df.end_time.apply(time_to_date)
    df["start_string"] = df.start_time.apply(lambda x: x.strftime("%I:%M"))
    df["end_string"] = df.end_time.apply(lambda x: x.strftime("%I:%M"))
    df["caption"] = (
        df["start_string"]
        + " to "
        + df["end_string"]
        + "<br>"
        + df["location"]
        + "<br>"
        + df["task_text"].replace("Disinfect & Prop Swap",
                                  "Disinfect<br>& Prop Swap")
    )
    df["class"] = df.task_text.apply(lambda r: task_classes.get(r, "unknown"))
    df["icons"] = df.task_text.apply(lambda r: icons.get(r))
    return df


def per_delta(start, end, delta):
    curr = start
    while curr < end:
        yield curr
        curr += delta
    # usage: for result in per_delta(datetime.datetime.now(), datetime.datetime.now().replace(hour=19))


def list_of_times(start, end, delta):
    return list(('"' + str(result.strftime("%-I:%M%p")) + '"') for result in per_delta(start, end, delta))


def day_of_times():
    return list_of_times(
        datetime.datetime(2020, 1, 1, 0, 0, 0, 0),
        datetime.datetime(2020, 1, 1, 23, 59, 59, 0),
        datetime.timedelta(minutes=15),
    )


long_locations = {
    "HTT": "Hit the Trail",
    "TH & PZ": "Toddlers Hollow & Putting Zoo",
    "P2P & STEMosphere": "Power2Play & STEMosphere",
    "Kid Grid & RPW": "Kid Grid & River Playway",
}


task_rows = {
    "Learn & Play": 2,
    "Closed to Public": 2,
    "Spray & Swap": 3,
    "Wipe": 4,
    "Swap": 3,
}

task_col_ranges = {
    "Learn & Play": (2, 3),
    "Closed to Public": (2, 3),
    "Spray & Swap": (3, 4),
    "Wipe": (4, 5),
    "Swap": (3, 5),
}

task_classes = {
    "Learn & Play": "learn_and_play",
    "Closed to Public": "closed_to_public",
    "Spray & Swap": "spray_and_swap",
    "Wipe": "wipe_down",
    "Swap": "prop_swap",
}

herb_list = [
                "Spray & Swap",
                "Swap",
                "Wipe",
            ]

icons = {
    "Learn & Play": ("close.png", "learn.png"),
    "Closed to Public": ("close.png", "learn.png"),
    "Spray & Swap": (
        "close.png",
        "swap.png",
        "spray.png",
    ),
    "Wipe": (
        "open.png",
        "wipe.png",
    ),
    "Swap": (
        "open.png",
        "swap.png",
    ),
}

weekdays = ["Monday", "Tuesday", "Wednesday",
            "Thursday", "Friday", "Saturday", "Sunday"]


# DAY_ROWS = [
#     (15, 20, "Monday"),  # Monday
#     (24, 28, "Tuesday"),  # Tuesday
#     (32, 39, "Wednesday"),  # Wednesday
#     (43, 50, "Thursday"),  # Thursday
#     (54, 61, "Friday"),  # Friday
#     (66, 73, "Weekend"),  # Weekend: Saturday & Sunday
# ]

DAY_ROWS = [
    (16, 24, "Monday"),  # Monday
    (29, 37, "Tuesday"),  # Tuesday
    (41, 49, "Wednesday"),  # Wednesday
    (53, 61, "Thursday"),  # Thursday
    (65, 73, "Friday"),  # Friday
    #(73, 80, "Weekend"),  # Weekend: Saturday & Sunday
    (77, 84, "Saturday"),
    (88, 95, "Sunday"),
]


def merged_size(cell, sheet):
    for rng in sheet.merged_cells:
        # print(cell.row, rng.left[0][0], ":", cell.column, rng.left[0][1])
        if cell.row == rng.left[0][0] and cell.column == rng.left[0][1]:
            return rng.size


if __name__ == "__main__":
    filename = sys.argv[1]
    date_range = sys.argv[2]
    print(f"Importing file: {filename}")
    wb = openpyxl.load_workbook(filename)
    sheet = wb["Schedule"]
    # row 12 is the first row of time headers, col B to AL
    task_list = []
    with open("tasks.csv", mode="w") as output_file:
        output_writer = csv.writer(
            output_file, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL
        )
        output_writer.writerow(
            ["weekday", "location", "start_time", "end_time", "task_text", "id"]
        )
        for day in DAY_ROWS:
            rng_start = "B" + str(day[0])
            rng_end = "AJ" + str(day[1])
            for row in sheet[rng_start:rng_end]:
                for cl in row:
                    if cl.value:
                        weekday = day[2]
                        location = sheet.cell(row=cl.row, column=1).value
                        start = sheet.cell(row=12, column=cl.column).value
                        mrg = merged_size(cl, sheet)
                        if mrg:
                            end = sheet.cell(
                                row=12, column=cl.column + mrg["columns"]
                            ).value
                        else:
                            end = sheet.cell(
                                row=12, column=cl.column + 1).value
                        task = cl.value
                        start_datetime = datetime.datetime.combine(datetime.date(2020, 1, 1), start)
                        end_datetime = datetime.datetime.combine(datetime.date(2020, 1, 1), end)
                        humanity_shift = {'start': datetime.datetime(2020, 1, 1, 12, 00, 00, 0), 'end': datetime.datetime(2020, 1, 1, 14, 15, 00, 0)}
                        if task in herb_list and ((start_datetime < humanity_shift['start']) or (humanity_shift['end'] <= start_datetime)):
                            assigned_to = "E"
                        else:
                            assigned_to = ""
                        id_placeholder = ""
                        output_writer.writerow(
                            [weekday, location, start, end, task, id_placeholder]
                        )
                        task_list.append(
                            {
                                'weekday': weekday,
                                'location': location,
                                'start': start, 'end': end,
                                'task': task,
                                'icons': icons.get(task),
                                'assigned_to': assigned_to,
                            })
    df = pd.DataFrame.from_records(task_list)
    df = df[~df['task'].str.contains("Members Only", na=False)]
    df = df[~df['task'].str.contains("Activity Room Play", na=False)]
    df = df[~df['task'].str.contains("Camp staff clean", na=False)]
    df = df[~df['task'].str.contains("Virex Spray Everything", na=False)]
    print(df)
    df["start_time"] = df.start.apply(
        lambda x: datetime.datetime.combine(datetime.date(
            2020, 1, 1), x)
    )
    df["end_time"] = df.end.apply(
        lambda x: datetime.datetime.combine(datetime.date(
            2020, 1, 1), x)
    )
    # df["start_string"] = df.start_time.apply(
    #     lambda x: '"' + x.strftime("%I:%M") + '"')
    # df["end_string"] = df.end_time.apply(
    #     lambda x: '"' + x.strftime("%I:%M") + '"')
    df["start_string"] = df.start_time.apply(
        lambda x: x.strftime("%-I:%M"))
    df["end_string"] = df.end_time.apply(
        lambda x: x.strftime("%-I:%M"))

    df["class"] = df.task.apply(lambda x: task_classes[x])
    df["long_location"] = df.location.apply(lambda x: long_locations.get(x))

    time_list = list_of_times(
        datetime.datetime.combine(datetime.date(
            2020, 1, 1), datetime.time(8, 0)),  # min(df['start'])),  # Earliest time in df
        datetime.datetime.combine(datetime.date(
            2020, 1, 1), datetime.time(18, 0)),  # max(df['end'])) + datetime.timedelta(minutes=15),  # Latest time in df
        datetime.timedelta(minutes=15),  # Interval of time list
    )
    hour_rows = [
        (index, time) for (index, time) in enumerate(time_list, 1) if ":00" in time
    ]
    # print("Hours:")
    print(hour_rows)

    df["start_row"] = df.start_time.apply(
        lambda x: time_list.index('"' + x.strftime("%-I:%M%p") + '"') + 1
    )
    df["end_row"] = df.end_time.apply(
        lambda x: time_list.index('"' + x.strftime("%-I:%M%p") + '"') + 1
    )
    df["time_range"] = df["start_string"] + " to " + df["end_string"]

    # df["start_row"] = df["start"]
    # df["end_row"] = df["end"]
    df["row"] = df.task.apply(lambda r: task_rows.get(r, 6))
    df["col_start"] = df.task.apply(lambda r: task_col_ranges.get(r, 6)[0])
    df["col_end"] = df.task.apply(lambda r: task_col_ranges.get(r, 7)[1])
    df = df.drop(['start_time', 'end_time', 'start', 'end'], axis=1)

    events = list(df.to_dict("index").values())
    #print(events)
    
    # The next line fails if there are days with no events!!!!!!!!!!!!!!!!
    #day_ranges = {
    #    wday: (min(df.loc[df['weekday'] == wday]['start']), max(df.loc[df['weekday'] == wday]['end'])) for wday in weekdays}
    #print("Day Ranges")
    #print(day_ranges)

    context = {
        "num_rows": len(time_list),
    #    "day_ranges": day_ranges,
        "date_range": date_range,
        # "weekday": weekday.lower().capitalize(),
        # "events": events,
        # "time_list": list(enumerate(time_list, 1)),
        # "hour_rows": hour_rows,
    }
    print(df.head())
    print(time_list)
    df.to_csv(r'_data/eventscsv.csv', index=False)

    with open(r'_data/context.yml', 'w') as yml_file:
        documents = yaml.dump(
            context, yml_file, default_flow_style=False)
    with open(r'_data/events.yml', 'w') as yml_file:
        documents = yaml.dump(
            events, yml_file, default_flow_style=False)
    with open(r'_data/time_list.yml', 'w') as yml_file:
        documents = yaml.dump(
            list(enumerate(time_list, 1)), yml_file, default_flow_style=False)
    with open(r'_data/hour_rows.yml', 'w') as yml_file:
        documents = yaml.dump(
            hour_rows, yml_file, default_flow_style=False)

    hard_prop_sked = pd.read_excel("Hard_Props_Pull_Schedule.xlsx", engine='openpyxl')
    hard_prop_sked.to_csv(r'_data/hard_prop.csv', index=False)
    hard_prop_sked_dict = list(hard_prop_sked.to_dict("index").values())
    print(hard_prop_sked)

    with open(r'_data/hard_props.yml', 'w') as yml_file:
        documents = yaml.dump(
            hard_prop_sked_dict, yml_file, default_flow_style=False)

